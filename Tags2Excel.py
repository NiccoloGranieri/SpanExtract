from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import re
import argparse
import os

parser = argparse.ArgumentParser()

parser.add_argument("--folderParsing", type=str, help="iterate through a folder of text files")
parser.add_argument("--sum", help="set the sum threshold", type=int)

args = parser.parse_args()

openFile = []

if args.folderParsing:
    folder = args.folderParsing
    print('Absolute path of the folder being parsed = ' + os.path.abspath(folder))
    for root, subdirs, files in os.walk(folder):
        for filename in files:
            if filename != '.DS_Store':
                file_path = os.path.join(root, filename)
                openFile.append(file_path)

if args.sum:
    threshold = args.sum
else:
    threshold = 3

def textLineParsing(file):
    fileLines = []
    with open (file, 'rt') as in_file:
        for line in in_file:
            fileLines.append(line)
    return fileLines

def lineDictionary(line):
    wordHits = dict()
    words = line.split()
    for word in words:
        if word in wordHits:
            wordHits[word] += 1
        else:
            wordHits[word] = 1
    return(wordHits)

def formatSheet(excelSheet, titles):
    cell = sheet1.cell(row = 1, column = 1, value = "No.")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 2, value = "Line No.")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 3, value = "File Name")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 4, value = "Speaker")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 5, value = "Line")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 6, value = "Line + 1")
    cell.font = Font(color = "FF0000", bold = True)
    for i in range(len(titles)):
        cell = sheet1.cell(row = 1, column = 7 + i, value = titles[i])
        cell.font = Font(color = "FF0000", bold = True)
        columnCount = 8 + i
    cell = sheet1.cell(row = 1, column = columnCount, value = "Sum")
    cell.font = Font(color = "FF0000", bold = True)
    return excelSheet

def writeToSheet(totalNumber, lineNumber, fileName, speaker, line, linePlusOne, features, instances, sum):
    cell = sheet1.cell(row = totalNumber + 1, column = 1, value = totalNumber)
    cell.font = Font(bold = True)
    cell = sheet1.cell(row = totalNumber + 1, column = 2, value = lineNumber)
    cell = sheet1.cell(row = totalNumber + 1, column = 3, value = fileName)
    cell = sheet1.cell(row = totalNumber + 1, column = 4, value = speaker)
    cell = sheet1.cell(row = totalNumber + 1, column = 5, value = line)
    cell = sheet1.cell(row = totalNumber + 1, column = 6, value = linePlusOne)
    for i in range(len(features)):
        cell = sheet1.cell(row = totalNumber + 1, column = 7 + i, value = instanceCount[features[i]])
        columnCount = 8 + i
    cell = sheet1.cell(row = totalNumber + 1, column = columnCount, value = sum)
    excelSheet.save(filename = "File Cluster.xlsx")

fileNumber = 0
features = ['we_#0', 'we_#1', 'think_#2', 'mean_#2', 'guess_#2', 'know_#2', '((laughs))', '((laughing))', '((chuckles))', '((chuckling))', '((hehe))', '((thh))', '((ehh))', '((heh))']
# features = ['((laughs))', '((laughing))', '((chuckles))', '((chuckling))', '((hehe))', '((thh))', '((ehh))', '((heh))']
totalLinesCount = 0
singleFileLinesCount = 0
openFile.sort()
hits = 0

if __name__ == '__main__':
    for inputFile in openFile:
        # If the file is the first file, create the workbook
        if fileNumber == 0:
            excelSheet = Workbook()
            sheet1 = excelSheet.active
            sheet1.title = 'Analysis'
            excelSheet = formatSheet(excelSheet, features)
        fileNumber += 1                                                                                                                                   
        print("Looking at file number " + str(fileNumber) + " in folder...") 
        fileName = os.path.basename(inputFile[:-4])
        fileLines = textLineParsing(inputFile)
        for i in range(len(fileLines)):
            lineNumber = i
            speaker = fileLines[i].partition(":")[0]
            count = 0
            instanceCount = {el:0 for el in features}
            for feature in features:
                if feature in lineDictionary(fileLines[i].lower()):
                    hits += 1
                    wordDict = lineDictionary(fileLines[i].lower())
                    instances = wordDict[feature]
                    instanceCount[feature] = instances
                    count = count + instances
            if count >= threshold:
                totalLinesCount += 1
                singleFileLinesCount += 1
                line = fileLines[i].partition(":")[2]
                nextLine = fileLines[i+1].partition(":")[2]
                totalCount = count
                print(instanceCount)
                writeToSheet(totalLinesCount, i, fileName, speaker, line, nextLine, features, instanceCount, totalCount)
        sheet1.merge_cells(start_row = (totalLinesCount + 2) - singleFileLinesCount, start_column = 3, end_row = totalLinesCount + 1, end_column = 3)
        mergedCell = sheet1.cell(row = (totalLinesCount + 2) - singleFileLinesCount, column = 3)
        mergedCell.alignment = Alignment(vertical='center')
        singleFileLinesCount = 0
        excelSheet.save(filename = "File Cluster.xlsx")
    print("I found " + str(hits) + " kinds of laughs")