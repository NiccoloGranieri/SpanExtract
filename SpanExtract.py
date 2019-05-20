import os
import sys
import re
import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

parser = argparse.ArgumentParser()

parser.add_argument("--fileParsing", type=str, help="load a single text file")
parser.add_argument("--folderParsing", type=str, help="iterate through a folder of text files")
parser.add_argument("--verbose", help="increase output verbosity", action="store_true")
parser.add_argument("--span", help="set the span of lines to extract", type=int)
parser.add_argument("--duplicates", help="sets the script to print double lines", action="store_true")
parser.add_argument("--mode", help="sets the mode of the script", type=int)
parser.add_argument("--tags", help="sets the tags to search in the text", type=str, nargs='*')
parser.add_argument("--featureTwo", help="sets the second feature to search around the tags", type=str, nargs='*')
parser.add_argument("--subMode", help="sets the submode of mode 1", type=int)
# parser.add_argument("--excel", help="enables excel output for mode 1", action="store_true")
parser.add_argument("--threshold", help="sets the threshold for the sum of the excel", type=int)

args = parser.parse_args()

hits = 0
secondFeatures = ['you']
openFile = []
fileOutput = 'txt'
lineCount = []
spaceSeeker = 0
printedSpans = 0
ignoredTags = 0
lastLinePrinted = -1
lastSpanPrinted = -1

totalLinesCount = 0
singleFileLinesCount = 0
threshold = 2

verboseprint = print if args.verbose else lambda *a, **k: None

if args.fileParsing:
    try:
        openfile = args.fileParsing
    except IndexError:
        print("Please provide a file name to parse.")
        sys.exit()

if args.folderParsing:
    folder = args.folderParsing
    verboseprint('Folder being parsed = ' + folder)
    print('Absolute path of the folder being parsed = ' + os.path.abspath(folder))
    for root, subdirs, files in os.walk(folder):
        for filename in files:
            if filename != '.DS_Store':
                file_path = os.path.join(root, filename)
                openFile.append(file_path)
    openFile.sort()

if args.span:
    span = args.span
else:
    span = 5

if args.duplicates:
    considerduplicates = True
else:
    considerduplicates = False

if args.mode:
    scriptMode = args.mode
else:
    scriptMode = 0

tags = ['((laughs))', '((laughing))', '((chuckles))', '((chuckling))', '((hehe))', '((heh))', '((ehh))', '((thh))']

if args.tags:
    tags.clear()
    for u in range (0, len(args.tags)):
        tags.append(args.tags[u])
    verboseprint("")
    verboseprint("Tags to look for: " + str(tags) + "\n""")

secondFeatures = ['((laughs))', '((laughing))', '((chuckles))', '((chuckling))', '((hehe))', '((heh))', '((ehh))', '((thh))']

if args.featureTwo:
    secondFeatures.clear()
    for p in range (0, len(args.featureTwo)):
        secondFeatures.append(args.featureTwo[p])
    verboseprint("Second Features: " + str(secondFeatures) + "\n""")

if args.subMode:
    submode = args.subMode
else:
    submode = 1.1

# if args.excel:
#     excel = True
# else:
#     excel = False

if args.threshold:
    threshold = args.threshold
else:
    threshold = 2
        
# def initOut(fileName, fileType, fileIndex):
#     # Initialises output file
#     if fileType == 'xls':
#         if fileIndex == 0:
#             excelSheet = Workbook()
#             sheet1 = excelSheet.active
#             sheet1.title = 'Analysis'
#             # excelSheet = formatSheet(excelSheet, features)
#             return excelSheet
#     elif fileType == 'txt':
#             textFile = open(fileName[:-4] + "_Parsed.txt", "w+")
#             return textFile

def textLineParsing(file):
    fileLines = []
    with open (file, 'rt') as in_file:
        for line in in_file:
            fileLines.append(line)
    return fileLines

def lineDictionary(line):
    wordHits = dict()
    words = line.split()
    for word in words:
        if word in wordHits:
            wordHits[word] += 1
        else:
            wordHits[word] = 1
    return(wordHits)

def identifySubmode(modeParser, returnType):
    if modeParser == 1:
        dictsToParse = [hitLineDict, lastLineDict]
        dictsToCount = [hitLineDict]
    elif modeParser == 2:
        dictsToParse = [hitLineDict, nextLineDict]
        dictsToCount = [hitLineDict]
    elif modeParser == 3:
        dictsToParse = [hitLineDict, nextLineDict, lastLineDict]
        dictsToCount = [hitLineDict, nextLineDict, lastLineDict]
    if (returnType == 'txt'): return dictsToParse
    else: return dictsToCount

def lookingForFeatureTwo(dictsToSearch, secondFeatures):
    for secondFeature in secondFeatures:
        for d in range(0, len(dictsToParse)):
            if secondFeature in dictsToParse[d]:
                verboseprint ("    I found " + secondFeature + " around line  " + str(i))
                return True
    return False

def formatSheet(excelSheet, titles):
    cell = sheet1.cell(row = 1, column = 1, value = "No.")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 2, value = "Line No.")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 3, value = "File Name")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 4, value = "Speaker")
    cell.font = Font(color = "FF0000", bold = True)
    cell = sheet1.cell(row = 1, column = 5, value = "Line")
    if submode == 3:
        cell.font = Font(color = "FF0000", bold = True)
        cell = sheet1.cell(row = 1, column = 6, value = "Line + 1")
        cell.font = Font(color = "FF0000", bold = True)
        cell = sheet1.cell(row = 1, column = 7, value = "Line + 2")
        lastCol = 7
    else:
        lastCol = 5
    cell.font = Font(color = "FF0000", bold = True)
    for i in range(len(titles)):
        cell = sheet1.cell(row = 1, column = (lastCol + 1) + i, value = titles[i])
        cell.font = Font(color = "FF0000", bold = True)
        columnCount = (lastCol + 1) + i
    cell = sheet1.cell(row = 1, column = columnCount + 1, value = "Sum")
    cell.font = Font(color = "FF0000", bold = True)
    return excelSheet

def writeToSheet(totalNumber, lineNumber, fileName, speaker, line, linePlusOne, linePlusTwo, features, instanceCount, sum, submode):
    cell = sheet1.cell(row = totalNumber + 1, column = 1, value = totalNumber)
    cell.font = Font(bold = True)
    cell = sheet1.cell(row = totalNumber + 1, column = 2, value = lineNumber)
    cell = sheet1.cell(row = totalNumber + 1, column = 3, value = fileName)
    cell = sheet1.cell(row = totalNumber + 1, column = 4, value = speaker)
    cell = sheet1.cell(row = totalNumber + 1, column = 5, value = line)
    if submode == 3:
        cell = sheet1.cell(row = totalNumber + 1, column = 6, value = linePlusOne)
        cell = sheet1.cell(row = totalNumber + 1, column = 7, value = linePlusTwo)
        lastCol = 7
    else:
        lastCol = 5
    for i in range(len(features)):
        cell = sheet1.cell(row = totalNumber + 1, column = (lastCol + 1) + i, value = instanceCount[features[i]])
        columnCount = (lastCol + 1) + i
    cell = sheet1.cell(row = totalNumber + 1, column = (columnCount + 1), value = sum)
    excelSheet.save(filename = "File Cluster.xlsx")

if __name__ == '__main__':
    for inputFile in openFile:
        verboseprint("Looking at file number " + str(inputFile) + " in folder...") 
        if scriptMode != 2: parsedFile = open(inputFile[:-4] + "_Parsed.txt", "w+")
        # fileName = os.path.basename(inputFile[:-4])
        fileLines = textLineParsing(inputFile)
        for i in range(len(fileLines)):
            hitLineDict = lineDictionary(fileLines[i].lower())
            for tag in tags:
                if tag in hitLineDict:
                    hits += hitLineDict[tag]
                    
                    # MODE 0
                    if scriptMode == 0:
                        verboseprint ("I found " + tag + " on line " + str(i))
                        min = i - span
                        max = i + span
                        if considerduplicates == False:
                            tagsInLine = hitLineDict[tag]
                        else:
                            tagsInLine = 1
                            ignoredTags += hitLineDict[tag] - tagsInLine
                        for k in range (tagsInLine):
                            verboseprint ("I found " + tag + " on line " + str(i))
                            for line in range(min, max + 1):
                                if line >= 0 and line < len(fileLines) and line not in lineCount:
                                    if considerduplicates:
                                        if line == min and lastLinePrinted < line and lineCount != []:
                                            parsedFile.write("\n")
                                        lineCount.append(line)
                                    parsedFile.write(str(line))
                                    parsedFile.write(" " + fileLines[line])
                                    lastLinePrinted = line
                            if considerduplicates == False:
                                parsedFile.write("\n")
                            printedSpans += 1
                                    
                    # MODE 1 INCLUDING SUBMODES
                    elif scriptMode == 1:
                        min = i - 1
                        max = i + 2
                        if i < len(fileLines) - 1: nextLineDict = lineDictionary(fileLines[i+1].lower())
                        if i < len(fileLines) - 2: lastLineDict = lineDictionary(fileLines[i+2].lower())
                        dictsToParse = identifySubmode(submode, 'txt')
                        tagsInLine = 1
                        ignoredTags += hitLineDict[tag] - tagsInLine
                        for k in range (tagsInLine):
                            verboseprint ("I found " + tag + " on line " + str(i))
                            if lookingForFeatureTwo(dictsToParse, secondFeatures):
                                if considerduplicates == True and i == lastSpanPrinted:
                                    pass
                                else:
                                    for line in range(min, max + 1):
                                        if line >= 0 and line < len(fileLines) and line not in lineCount:
                                            parsedFile.write(str(line))
                                            parsedFile.write(" " + fileLines[line])
                                    parsedFile.write("\n")
                                    lastSpanPrinted = i
                                    printedSpans += 1

        if scriptMode == 2:
            if inputFile == openFile[0]:
                excelSheet = Workbook()
                sheet1 = excelSheet.active
                sheet1.title = 'Analysis'
                excelSheet = formatSheet(excelSheet, tags)                                                                                                                              
            fileName = os.path.basename(inputFile[:-4])
            for i in range(len(fileLines)):
                hitLineDict = lineDictionary(fileLines[i].lower())
                if i < len(fileLines) - 1: nextLineDict = lineDictionary(fileLines[i+1].lower())
                if i < len(fileLines) - 2: lastLineDict = lineDictionary(fileLines[i+2].lower())
                dictsToCount = identifySubmode(submode, 'xls')
                lineNumber = i
                speaker = fileLines[i].partition(":")[0]
                instanceCount = {el:0 for el in tags}
                if any(i in hitLineDict for i in tags):
                    for d in range(0, len(dictsToCount)):
                        wordDict = dictsToCount[d]
                        for feature in tags:
                            if feature in wordDict: 
                                instanceCount[feature] += wordDict[feature]
                count = sum(instanceCount.values())
                if count >= threshold:
                    totalLinesCount += 1
                    singleFileLinesCount += 1
                    line = fileLines[i].partition(":")[2]
                    if i < len(fileLines) - 1: nextLine = fileLines[i+1].partition(":")[2]
                    if i < len(fileLines) - 2: lastLine = fileLines[i+2].partition(":")[2]
                    verboseprint(instanceCount)
                    writeToSheet(totalLinesCount, i, fileName, speaker, line, nextLine, lastLine, tags, instanceCount, count, submode)
            startRowMerge = (totalLinesCount + 2) - singleFileLinesCount
            endRowMerge = totalLinesCount + 1
            if startRowMerge < endRowMerge:
                sheet1.merge_cells(start_row = (totalLinesCount + 2) - singleFileLinesCount, start_column = 3, end_row = totalLinesCount + 1, end_column = 3)
                mergedCell = sheet1.cell(row = (totalLinesCount + 2) - singleFileLinesCount, column = 3)
                mergedCell.alignment = Alignment(vertical='center')
            singleFileLinesCount = 0
            excelSheet.save(filename = "File Cluster.xlsx")
    verboseprint("I found " + str(hits) + " tags in total.")
    verboseprint("I printed " + str(printedSpans) + " spans in the text file.")
    verboseprint("I ignored " + str(ignoredTags) + " tags in total.")
    if scriptMode != 2: parsedFile.close()